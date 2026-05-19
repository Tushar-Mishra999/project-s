"""PDF / DOCX / XLSX rendering — replaces Puppeteer / html-to-docx / SheetJS."""
import io
import re
from bs4 import BeautifulSoup

_FENCE_RE = re.compile(r'^```(?:html)?\s*|\s*```$', re.MULTILINE)

def _strip_fences(html: str) -> str:
    return _FENCE_RE.sub('', html).strip()

_PDF_CSS = """<style>
div,section,article,aside,header,footer,main,nav,p,li,ul,ol,span,blockquote {
  border: none !important;
  box-shadow: none !important;
  border-radius: 0 !important;
  background: transparent !important;
  padding: 0 !important;
  margin-bottom: 4px !important;
}
h1,h2,h3,h4,h5,h6 {
  border: none !important;
  box-shadow: none !important;
  border-radius: 0 !important;
  background: transparent !important;
  padding: 2px 0 !important;
  margin-top: 14px !important;
  margin-bottom: 4px !important;
}
body { font-family: Helvetica, Arial, sans-serif; font-size: 11px; }
table { width: 100%; margin-bottom: 12px; }
th, td { border: 1px solid #ccc; padding: 4px 6px; font-size: 10px; }
th { background-color: #1e3a5f; color: white; font-weight: bold; }
</style>"""

# CSS properties that cause xhtml2pdf table row-height crashes
_TABLE_STYLE_STRIP = re.compile(
    r'border-collapse\s*:[^;]+;?\s*|'
    r'border-radius\s*:[^;]+;?\s*|'
    r'box-shadow\s*:[^;]+;?\s*',
    re.IGNORECASE,
)

def _sanitize_tables(html: str) -> str:
    """Remove inline styles on tables/rows that crash xhtml2pdf, and drop colgroup/col."""
    soup = BeautifulSoup(html, "lxml")
    for tag in soup.find_all(["colgroup", "col"]):
        tag.decompose()
    for table in soup.find_all("table"):
        if table.get("style"):
            table["style"] = _TABLE_STYLE_STRIP.sub("", table["style"]).strip().rstrip(";")
    for tr in soup.find_all("tr"):
        style = tr.get("style", "")
        # Remove height from tr — causes row-height count mismatch in xhtml2pdf
        style = re.sub(r'height\s*:[^;]+;?\s*', '', style, flags=re.IGNORECASE)
        if style.strip():
            tr["style"] = style
        elif "style" in tr.attrs:
            del tr["style"]
    # Return just the body content (lxml wraps in html/body)
    body = soup.find("body")
    return "".join(str(c) for c in body.children) if body else html


def render_pdf(html: str) -> bytes:
    from xhtml2pdf import pisa
    html = _strip_fences(html)
    html = _sanitize_tables(html)
    if "</head>" in html:
        html = html.replace("</head>", _PDF_CSS + "</head>", 1)
    else:
        html = _PDF_CSS + html
    buf = io.BytesIO()
    pisa.CreatePDF(html, dest=buf)
    return buf.getvalue()


def render_docx(html: str) -> bytes:
    from docx import Document
    from htmldocx import HtmlToDocx
    html = _strip_fences(html)
    doc = Document()
    HtmlToDocx().add_html_to_document(html, doc)
    buf = io.BytesIO()
    doc.save(buf)
    return buf.getvalue()


def render_xlsx(html: str) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    soup = BeautifulSoup(html, "lxml")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    tables = soup.find_all("table")
    if not tables:
        ws = wb.create_sheet("Report")
        for i, elem in enumerate(soup.find_all(["h1", "h2", "h3", "p", "li"]), 1):
            ws.cell(row=i, column=1, value=elem.get_text(strip=True))
    else:
        for t_idx, table in enumerate(tables):
            prev_heading = table.find_previous(["h1", "h2", "h3"])
            sheet_name = (prev_heading.get_text(strip=True)[:31] if prev_heading else f"Sheet{t_idx + 1}")
            ws = wb.create_sheet(title=sheet_name)
            header_fill = PatternFill("solid", fgColor="1E293B")
            header_font = Font(color="FFFFFF", bold=True, size=10)
            for r_idx, row in enumerate(table.find_all("tr"), 1):
                cells = row.find_all(["th", "td"])
                for c_idx, cell in enumerate(cells, 1):
                    val = cell.get_text(strip=True)
                    ws_cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    if cell.name == "th":
                        ws_cell.fill = header_fill
                        ws_cell.font = header_font
                    ws_cell.alignment = Alignment(wrap_text=True)
                    # Auto-width approximation
                    col_letter = ws_cell.column_letter
                    current = ws.column_dimensions[col_letter].width
                    ws.column_dimensions[col_letter].width = max(current, min(len(val) + 2, 40))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
